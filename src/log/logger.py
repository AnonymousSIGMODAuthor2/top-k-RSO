import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExperimentLogger:
    def __init__(self, experiment_name, baseline_name="base_iadu"):
        """
        experiment_name: Name of the output file.
        baseline_name: The prefix of the method to use as the baseline for Diff% and Speedup calculations.
                       Default is 'base_iadu'. For no-rF experiments, set this to 'base_iadu_no_r'.
        """
        self.filename = f"{experiment_name}.xlsx"
        self.baseline_name = baseline_name
        self.logs = []
        
        # Standard metadata columns that should always appear first
        self.meta_columns = ["shape", "K", "k", "W", "g*K/k", "G", "lenCL"]
        
        # The standard metrics we expect for every method
        self.metric_suffixes = ["_hpfr", "_pss_sum", "_psr_sum"]
        self.time_suffixes = ["_prep_time", "_sel_time", "_x_time"]

    def log(self, row_dict):
        """
        Logs a dictionary of results.
        Keys should follow the format: '{method}_{metric}' (e.g., 'grid_weighted_hpfr')
        """
        self.logs.append(row_dict)

    def _calculate_derived_metrics(self, df):
        """
        Dynamically calculates Diff% and Speedup relative to the specified baseline_name.
        """
        base_hpfr = f"{self.baseline_name}_hpfr"
        base_time = f"{self.baseline_name}_x_time"
        
        # 1. Identify all unique methods present in the columns
        methods = set()
        for col in df.columns:
            # We check for score columns to identify methods
            # Must exclude the baseline itself so we don't diff against self
            if col.endswith("_hpfr") and col != base_hpfr:
                methods.add(col.replace("_hpfr", ""))
        
        # 2. Calculate Metrics for each identified method
        for method in methods:
            # A. HPFR Diff % (Method vs Base)
            score_col = f"{method}_hpfr"
            diff_col = f"{method}_hpfr_diff%"
            
            if score_col in df.columns and base_hpfr in df.columns:
                df[diff_col] = df.apply(
                    lambda row: (row[score_col] - row[base_hpfr]) / row[base_hpfr] if row[base_hpfr] != 0 else 0,
                    axis=1
                )

            # B. Speedup (Base Time / Method Time)
            time_col = f"{method}_x_time"
            speedup_col = f"{method}_speedup"
            
            if time_col in df.columns and base_time in df.columns:
                df[speedup_col] = df.apply(
                    lambda row: row[base_time] / row[time_col] if row[time_col] != 0 else 0,
                    axis=1
                )
                
        return df

    def _organize_columns(self, df):
        """
        Sorts columns logically: Meta -> Baseline -> Methods (Alphabetical) -> Times
        """
        cols = list(df.columns)
        
        # 1. Metadata (Fixed order)
        final_cols = [c for c in self.meta_columns if c in cols]
        
        # 2. Baseline Metrics (Always second, based on configured baseline_name)
        # Excludes time/speedup columns which go to the end
        base_metrics = [c for c in cols if c.startswith(f"{self.baseline_name}_") and not any(t in c for t in ["time", "speedup"])]
        final_cols.extend(sorted(base_metrics))
        
        # 3. Identify Other Methods
        methods = set()
        for c in cols:
            for suffix in self.metric_suffixes:
                if c.endswith(suffix) and self.baseline_name not in c:
                    methods.add(c.replace(suffix, ""))
        
        sorted_methods = sorted(list(methods))
        
        # 4. Method Metrics (Score, Diff, PSS, PSR)
        for method in sorted_methods:
            standard_block = [
                f"{method}_hpfr",
                f"{method}_hpfr_diff%",
                f"{method}_pss_sum",
                f"{method}_psr_sum"
            ]
            for col in standard_block:
                if col in cols:
                    final_cols.append(col)
                    
        # 5. Times & Speedup (At the end)
        
        # Baseline Time
        base_times = [c for c in cols if c.startswith(f"{self.baseline_name}_") and "time" in c]
        final_cols.extend(sorted(base_times))
        
        # Method Times
        for method in sorted_methods:
            time_block = [
                f"{method}_prep_time",
                f"{method}_sel_time",
                f"{method}_x_time",
                f"{method}_speedup"
            ]
            for col in time_block:
                if col in cols:
                    final_cols.append(col)
                    
        # 6. Catch-all for anything missed
        existing_set = set(final_cols)
        remaining = [c for c in cols if c not in existing_set]
        final_cols.extend(remaining)
        
        return df[final_cols]

    def save(self):
        if not self.logs:
            print("No logs to save.")
            return

        # 1. Process Detailed Data
        df = pd.DataFrame(self.logs)
        df = self._calculate_derived_metrics(df)
        df_detailed = self._organize_columns(df)

        # 2. Create Summary (Group by Settings)
        group_keys = [k for k in ["K", "k", "W", "g*K/k", "G"] if k in df_detailed.columns]
        
        if group_keys:
            df_summary = df_detailed.groupby(group_keys).mean(numeric_only=True).reset_index()
            # Recalculate metrics (Diff%, Speedup) on the averages to be mathematically correct
            df_summary = self._calculate_derived_metrics(df_summary)
            df_summary = self._organize_columns(df_summary)
            
            # Drop meaningless columns from summary
            drop_cols = [c for c in ["shape", "lenCL"] if c in df_summary.columns]
            df_summary = df_summary.drop(columns=drop_cols)
        else:
            df_summary = pd.DataFrame()

        # 3. Save to Excel
        try:
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                if not df_summary.empty:
                    df_summary.to_excel(writer, sheet_name='Settings Summary', index=False)
                df_detailed.to_excel(writer, sheet_name='Detailed Results', index=False)
            
            self._apply_pro_styling()
            print(f"✔ Results saved correctly to: {self.filename}")
        except PermissionError:
            print(f"❌ ERROR: Could not save to '{self.filename}'. Is the file open in Excel?")
        except Exception as e:
            print(f"❌ ERROR Saving Log: {e}")

    def _apply_pro_styling(self):
        try:
            wb = load_workbook(self.filename)
            
            colors = {
                "header": "404040", "meta": "E7E6E6", "base": "FBE5D6", 
                "score": "E2EFDA", "diff": "D0CECE", "time": "DDEBF7", "speedup": "C6E0B4"
            }
            
            for ws in wb.worksheets:
                # Headers
                header_fill = PatternFill(start_color=colors["header"], fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                ws.freeze_panes = "B2"

                # Columns
                for i, col_cells in enumerate(ws.columns, start=1):
                    header_val = str(col_cells[0].value)
                    col_letter = get_column_letter(i)
                    
                    # Width
                    max_len = max([len(str(c.value)) if c.value is not None else 0 for c in col_cells[:15]])
                    ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

                    # Colors
                    fill_color = None
                    if any(x in header_val for x in self.meta_columns): fill_color = colors["meta"]
                    # Dynamic check for baseline name
                    elif self.baseline_name in header_val: fill_color = colors["base"]
                    elif "diff%" in header_val: fill_color = colors["diff"]
                    elif "speedup" in header_val: fill_color = colors["speedup"]
                    elif "time" in header_val: fill_color = colors["time"]
                    elif any(x in header_val for x in ["hpfr", "pss", "psr"]): fill_color = colors["score"]

                    if fill_color:
                        fill_obj = PatternFill(start_color=fill_color, fill_type="solid")
                        border = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"))
                        for cell in col_cells[1:]:
                            cell.fill = fill_obj
                            cell.border = border
                            
                            # Number Formatting
                            if isinstance(cell.value, (int, float)):
                                if "diff%" in header_val: cell.number_format = '0.00%'
                                elif "speedup" in header_val: cell.number_format = '0.00"x"'
                                elif "time" in header_val: cell.number_format = '0.000'
                                elif "hpfr" in header_val: cell.number_format = '0.000'
                                elif abs(cell.value) < 0.01: cell.number_format = '0.00E+00'

            wb.save(self.filename)
        except Exception as e:
            print(f"Warning: Styling step failed: {e}")